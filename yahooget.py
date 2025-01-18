import os
from yahoo_fin.stock_info import get_data, tickers_sp500, tickers_nasdaq, tickers_other, get_quote_table, get_income_statement
import yfinance as yf
import openpyxl
import time

# Create a new workbook
wb = openpyxl.Workbook()

def get_historical_data(ticker, start_date=1/1/2000, end_date=None):
    """
    Fetches historical stock data for the given ticker and prints it.
    """
    try:
        # Fetch historical data
        datamess = get_data(ticker, start_date=start_date, end_date=end_date)
        print("\nHistorical Data:")
        print(datamess)
        return datamess
    except Exception as e:
        print(f"Error fetching historical data for {ticker}: {e}")

def income_statement_with_yfinance(ticker):
    """
    Fetches the income statement using the yfinance library and prints it.
    """
    try:
        stock = yf.Ticker(ticker)
        financials = stock.financials  # Income statement
        if financials.empty:
            print(f"No income statement data available for {ticker}.")
        else:
            print(f"\nIncome Statement for {ticker}:")
            print(financials)
        return financials
    except Exception as e:
        print(f"Error fetching income statement for {ticker} with yfinance: {e}")

def balance_sheet_with_yfinance(ticker):
    """
    Fetches the balance sheet using the yfinance library and prints it.
    """
    try:
        stock = yf.Ticker(ticker)
        balance_sheet = stock.balance_sheet  # Balance sheet
        if balance_sheet.empty:
            print(f"No balance sheet data available for {ticker}.")
        else:
            print(f"\nBalance Sheet for {ticker}:")
            print(balance_sheet)
        return balance_sheet
    except Exception as e:
        print(f"Error fetching balance sheet for {ticker} with yfinance: {e}")

def cash_flow_statement_with_yfinance(ticker):
    """
    Fetches the cash flow statement using the yfinance library and prints it.
    """
    try:
        stock = yf.Ticker(ticker)
        cashflow = stock.cashflow  # Cash flow statement
        if cashflow.empty:
            print(f"No cash flow statement data available for {ticker}.")
        else:
            print(f"\nCash Flow Statement for {ticker}:")
            print(cashflow)
        return cashflow
    except Exception as e:
        print(f"Error fetching cash flow statement for {ticker} with yfinance: {e}")

def adjust_column_width(sheet):
    """
    Adjusts the column width of all columns in the sheet based on the longest string in each column.
    For financial statements, it checks both columns individually.
    """
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter  # Get the column name (e.g. 'A', 'B', 'C', ...)
        
        for cell in column:
            try:
                # Get the length of the string representation of each cell and keep track of the max length
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Adjust the column width, add extra space for padding
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width


def write_to_workbook(ticker, hist_data, income_statement, balance_sheet, cash_flow):
    """
    Writes the fetched data to a new Excel workbook with auto-adjusted column widths.
    """
    file_path = f"sheets/{ticker}_stock_data.xlsx"
    
    # Check if the file already exists
    if os.path.exists(file_path):
        overwrite = input(f"The file {ticker}_stock_data.xlsx already exists. Would you like to overwrite it? (Y/N): ").strip()
        if overwrite.upper() != 'Y':
            print(f"Skipping the overwrite. Data for {ticker} will not be saved.")
            return

    # Create a new workbook if not exists
    wb = openpyxl.Workbook()

    # Write historical data
    sheet = wb.create_sheet(title=f"{ticker} - Historical Data")
    sheet.append(["Date", "Open", "High", "Low", "Close", "Volume"])
    for index, row in hist_data.iterrows():
        sheet.append([index.date(), row["open"], row["high"], row["low"], row["close"], row["volume"]])

    # Adjust column widths for historical data sheet
    adjust_column_width(sheet)

    # Write income statement
    sheet = wb.create_sheet(title=f"{ticker} - Income Statement")
    sheet.append(["Metric", "Value"])
    for column in income_statement.columns:
        for key, value in income_statement[column].to_dict().items():
            # Convert dictionary or NaN values
            if isinstance(value, dict):
                value = str(value)  # Convert the dictionary to string
            elif isinstance(value, float) and (value != value):  # Check for NaN
                value = None  # Replace NaN with None
            sheet.append([key, value])

    # Adjust column widths for income statement sheet
    adjust_column_width(sheet)

    # Write balance sheet
    sheet = wb.create_sheet(title=f"{ticker} - Balance Sheet")
    sheet.append(["Metric", "Value"])
    for column in balance_sheet.columns:
        for key, value in balance_sheet[column].to_dict().items():
            # Convert dictionary or NaN values
            if isinstance(value, dict):
                value = str(value)  # Convert the dictionary to string
            elif isinstance(value, float) and (value != value):  # Check for NaN
                value = None  # Replace NaN with None
            sheet.append([key, value])

    # Adjust column widths for balance sheet sheet
    adjust_column_width(sheet)

    # Write cash flow statement
    sheet = wb.create_sheet(title=f"{ticker} - Cash Flow Statement")
    sheet.append(["Metric", "Value"])
    for column in cash_flow.columns:
        for key, value in cash_flow[column].to_dict().items():
            # Convert dictionary or NaN values
            if isinstance(value, dict):
                value = str(value)  # Convert the dictionary to string
            elif isinstance(value, float) and (value != value):  # Check for NaN
                value = None  # Replace NaN with None
            sheet.append([key, value])

    # Adjust column widths for cash flow statement sheet
    adjust_column_width(sheet)

    # Save the workbook
    wb.save(file_path)
    print(f"\nData for {ticker} has been written to {file_path}")



def main():
    """
    Main program loop for fetching stock data and confirming writing to workbook.
    """
    q = 'Y'
    while q.upper() == 'Y':
        entry = input("Enter a stock ticker to fetch data for: ")

        try:
            # Fetch and display historical data
            print("\nFetching historical data...\n")
            hist_data = get_historical_data(entry)

            # Fetch and display income statement
            print("\nFetching income statement...\n")
            income_statement = income_statement_with_yfinance(entry)

            # Fetch and display balance sheet
            print("\nFetching balance sheet...\n")
            balance_sheet = balance_sheet_with_yfinance(entry)

            # Fetch and display cash flow statement
            print("\nFetching cash flow statement...\n")
            cash_flow = cash_flow_statement_with_yfinance(entry)

            # Ask user if they want to write the data to a new workbook
            q_write = input(f"\nWould you like to write the data for {entry} to a new workbook? (Y/N): ").strip()
            if q_write.upper() == 'Y':
                write_to_workbook(entry, hist_data, income_statement, balance_sheet, cash_flow)
            
        except KeyError:
            print("Ticker does not exist. Please try again.")
        except Exception as e:
            print(f"An unexpected error occurred: {e}")

        # Ask the user if they want to fetch data for another stock
        q = input("Would you like to enter another stock? (Y/N): ").strip()

if __name__ == '__main__':
    main()
