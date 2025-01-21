import os
import openpyxl
import datetime
from openpyxl.styles import NamedStyle
from yahooget import get_stock_info  # Import get_stock_info from yahooget.py
from pathlib import Path


def create_dollar_style():
    """Creates a NamedStyle for dollar formatting."""
    return NamedStyle(name="dollar_style", number_format='"$"#,##0.00')


def create_percent_style():
    """Creates a NamedStyle for percentage formatting."""
    return NamedStyle(name="percent_style", number_format="0.0%")


def adjust_column_width(sheet):
    """Adjusts column widths based on the content."""
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_name].width = adjusted_width


def write_to_workbook(ticker, hist_data, income_statement, balance_sheet, cash_flow):
    """Writes stock data and financial statements to a new Excel file."""
    os.makedirs("sheets", exist_ok=True)
    now = datetime.datetime.now()
    format_now = now.ctime()
    sheets_dir = f"sheets/{ticker}_stock_data.xlsx"
    file_path = Path(sheets_dir)

    if os.path.exists(file_path):
        overwrite = input(
            f"The file {ticker}_stock_data.xlsx already exists. "
            "Would you like to overwrite it? (Y/N): "
        ).strip()
        if overwrite.upper() != "Y":
            print(f"Skipping the overwrite. Data for {ticker} will not be saved.")
            return

    wb = openpyxl.Workbook()

    # Remove the default "Sheet" tab if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Fetch stock info and create "Stock Info" sheet
    stock_info = get_stock_info(ticker)
    sheet = wb.create_sheet(title=f"{ticker} - Stock Info")
    sheet.append(["Report created:", format_now])  # Adds created date

    for key, value in stock_info.items():
        sheet.append([key, value])

    adjust_column_width(sheet)

    # Create and populate "Historical Data" sheet
    sheet = wb.create_sheet(title=f"{ticker} - Historical Data")
    sheet.append(["Date", "Open", "High", "Low", "Close", "Volume"])
    for index, row in hist_data.iloc[::-1].iterrows():
        sheet.append(
            [index.date(), row["open"], row["high"], row["low"], row["close"], row["volume"]]
        )

    # Apply dollar formatting to numerical columns (Open, High, Low, Close)
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.00'

    adjust_column_width(sheet)

    # Create and populate financial statements sheets
    for name, data in [
        ("Income Statement", income_statement),
        ("Balance Sheet", balance_sheet),
        ("Cash Flow Statement", cash_flow),
    ]:
        sheet = wb.create_sheet(title=f"{ticker} - {name}")
        reversed_data = data.iloc[:, ::-1]
        headers = ["Metric"] + list(reversed_data.columns) + ["Avg Yearly Change"]
        sheet.append(headers)

        for row_label in reversed_data.index:
            values = reversed_data.loc[row_label].tolist()
            try:
                percent_changes = reversed_data.loc[row_label].pct_change(
                    fill_method=None
                ).dropna()
                avg_change = percent_changes.mean() if not percent_changes.empty else None
            except ZeroDivisionError:
                avg_change = None

            sheet.append([row_label] + values + [avg_change])

            # Apply dollar formatting to numeric columns
            for col in range(1, len(values) + 1):
                if isinstance(values[col - 1], (int, float)):
                    sheet.cell(row=sheet.max_row, column=col + 1).number_format = '"$"#,##0.00'

            # Apply percentage formatting for average change
            avg_change_cell = sheet.cell(row=sheet.max_row, column=len(values) + 2)
            if avg_change is not None:
                avg_change_cell.value = avg_change
                avg_change_cell.number_format = "0.0%"

        adjust_column_width(sheet)

    wb.save(file_path)
    print(f"Data for {ticker} has been written to {file_path}")
